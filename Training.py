#Working with variables

# str="Pradyumna"
# num=25
# num1=35.67
#
#
# str="wwww.apple.com"
#
# num=45
# num1=55.67
#
#
# str="pradyumna"
# str="c"
#
# num=45

# Input and Output
#
# str="Pradyumna"
# print(str)
#
# num=45
# print(num)
#
# num1=45.50
# print(num1)
#
# str="I like to have Coffee"
# print(str)

#Formatting
#
# print(1,2,3,4,5,sep="*")
# #
# # name="Pradyumna"
# # company="DXC"
# #
# # print("My name is {},company is {}".format(name,company))
#
#
# print("My name is {name},and i go to {school}".format(name="Samrudh",school="NPS"))


# # Input
# name=input("Please Enter the name :")
# company=input("Please enter the company :")
# print("My name is{name} and i work in {company}".format(name=name,company=company))

#Working with Operators

# Arithamtic Operators

# x=25
# y=15
#
# print(x+y)
# print(x-y)
# print(x%y)
# print(x//y)
# print(x*y)
# print(x/y)
#
# # Comparison Operators
#
# print(x<y)
# print(x>y)
# print(x==y)
# print(x>=y)
# print(x<=y)
# print(x!=y)
#
# #assignment Operators
# x=1
# y=1
# x=x+1,
# x+=1

# #Logical Operators
#
# x=True
# y=False
#
# print(x and y)
# print(x or y )
# print( not x)

#Identity Opearators
# is , is not

# x=5
# y=10
# print( x is y )
#
# print(id(x))
# print(id(y))

# Membership Operators

# in ,not in

# str="Pradyumna"
#
# print('z' in str)
# print('z' not in str)

# Type Casting
# str="1"
# num=str(str)
#
# print(type(str))
# print(type(num))
#
# num="11"
# num1="25"
#
# result=int(num)+int(num1)
# print(result)

# num=float(input("Enter the First Number :"))
# num1=float(input("Enter the Second Number :"))
#
# result=num+num1
# print(result)

# Strings Functions
#
# str="hello world"
# print(str.capitalize())
#
# print(str.lower())
# print(str.upper())
#
# print(str.count('l'))
# print(str.find('e'))
#
# # print(str[str.find('e'):5])
# print(str.find('l'))
#
# print(str.split(' '))
# print(str.replace("hello","Miss"))
# print(str)
# # print(str1)

# String Slicing

# str="Pradyumna"
#
# print(str[5])
#
# # print(str[start:Stop:step])
# print(str[0:6:2])
#
# print(str[5::-1])
#
# str1=str[-1::-1]
# print(str==str1)

# Conditional Statements

# if <condition>:
#     print("Something")
#
#
# if <condition1>:
#     print("something")
# else:
#     print("Condition 2")
#
#
# if <condition1>:
#     print("Something")
# elif <condition2>:
#     print("Something")
# else:
#     print("Soemthing")

# sex="aaa"
#
# if sex=='m':
#     print("Male Gender")
# elif sex=='f':
#     print("Female Gender")
# else:
#     print("Gender Unknown")
#
#
# #Given number id odd or even
#
# num=float(input("Enter the Number :"))
# if num>0:
#     if num%2==0:
#         print("Even")
#     else:
#         print("Odd")
# else:
#     print("Enter a number greater than zero")

# List -> Data Structure

# my_list=[1,2,3,4.5,5,6]
# my_list1=["1","2","3","4","5"]
#
# print(my_list)
# print(my_list1)
#
# print(type(my_list))
# print(type(my_list1))

# courses=["Maths","Science","Biology","Botany","Socials"]
# print(courses)
# courses.append("English")
# print(courses)
#
# courses1=["Sanskrit","Hindi"]
# courses.extend(courses1)
# print(courses)
#
# print(courses.index("Biology"))
# courses.insert(2,"Commerce")
# print(courses)
#
# courses.remove("Commerce")
# print(courses)
# # courses.remove(2) # Value Error not in list
#
# popped=courses.pop()
# print(courses)
# print(popped)
#
# # courses.clear()
#
# print(courses)
#
# # str='Hello\'s World'
#
# # Length of the List
# print(len(courses))

# courses=["Maths","Science","Biology","Botany","Socials"]
# courses.sort()
# # print(courses)
# # print(courses)
#
# sort_courses=sorted(courses)
# print(sort_courses)
# print(courses)
#
# courses.sort(reverse=True)
# print(courses)
#
# # del courses
# # sort_courses=sorted(courses,reverse=True)
# print(courses)
#
# num=[9,1,3,5,7,6]
# # num.sort(reverse=True)
# num1=sorted(num,reverse=True)
# print(num1)

# Looping
#
# while (Condition):
#     code

# i=1
# n=25
# while i<=n:
#     print(i)
#     # i=i+1
#     i+=1

# for item in something:
#     code
# for i in range(0,26):
#     print(i)

# for i in range(5,25):
#     print(i)

# for i in range(0,25,2):
#     print(i)

# for i in range(26,1,-1):
#     print(i)

# for i in range(26):
#     if i ==15:
#         continue
#     print(i)

# for i in range(26):
# #     if i ==15:
# #         break
# #     print(i)

# courses=["Maths","Science","Biology","Botany","Socials"]
# for course in courses:
#     print(course)

# for i in a:
#     print(i)
#
# i=1
# n=10
# while i<=n:
#     print(i)
#     i=i+1

# a=[9,5,4,69999999999,2,7,3]
# max_num=a[0]
# for i in a:
#     if i>max_num:
#         max_num=i
#
# print(max_num)

#
# #Linear Search
# a=[3,6,9,899,11,44,77,899,101]
# x=int(input("Enter the number :"))
# # for i in a:
#     if i==x:
#         print(a.index(i))

#enumerate

# for index,i in enumerate(a,start=1):
#     if i==x:
#         print(index)

# length=len(a)
# # for i in range(length):
# #     if a[i]==x:
# #         print(i)


# odd and Even number
#
# odd=[]
# even=[]
# for i in range(51):
#     if i%2==0:
#         even.append(i)
#     else:
#         odd.append(i)
#
# # print("Odd List :"+odd)
# print("Even List :",even)

# a=[9,4,8,3,2,1]
# b=a
# # # b=a.copy()
# # b=a
# # print(id(b))
# # print(id(a))
# b[1]=5
# print(a)
# print(b)

# Working with tuples
#
# my_tuple=(1,2,3,4,5,6,7,8)
# # my_tuple[2]=45
#
# a=1,2,"crew","facet","facet"
# print(a)
#
# # Tuple Unpacking
# b,c,d,e,f=a
# print(b)
# print(c)
# print(d)
# print(e)
#
# # Indexing
#
# print(a.index("crew"))
# print(a.count("facet"))
#
# # Slicing
#
# print(a[0:2])
#
# # iterate over a Tuple
#
# for i in a:
#     print(i)

# b=(1,2,3,[4,5],[6,25])
# b[3][0]=25
# print(b)
# # a=[1,2,3,(4,5)]
# for i in b:
#     if type(i)==list:
#         print(i.index(25))

# To print nested values
# a=[[1,2,3],[4,5,6],[7,8,9]]
# b=([1,2,3],[4,5,6],[7,8,9])
#
# for i in a:
#     for j in i:
#         print(j)
#
# for i in b:
#     for j in i:
#         print(j)

# Working with Sets

# a={"PRady","divakar","manoj","magu"}
# b={"sandeep","avinash","mahesh","prashanth"}
# a=list([1,2,3,4])
# b=tuple((1,2,3,4))
# c=set({1,2,3,4})
# a=list([1,2,3,4])
# print(a)
# #
# print(a)
# print(a)
# print(a)
# print(a)
# print(a)
#
# #Adding Elements to set
# a.update({"deepak","kiran"})
# a.update({"mooshandi"})
#
# print(a)
#
#
# # to remove an element from set
# a.remove("mooshandi")
# print(a)
#
# # Pop
# b=a.pop()
# print(b)

# a={"PRady","divakar","manoj","magu"}
# b={"sandeep","avinash","mahesh","prashanth","divakar"}
# # a.add("someting")
#
# union_set=a.union(b)
# print(union_set)
# print(a.intersection(b))
# print(a.difference(b))
# print(a.symmetric_difference(b))

# for item in a:
#     print(item)

# # Set should have only Immutable Elements . It cannot contain "Mutable elements"
# b={1,2,'abc',(4,5,6),[1,2,3]}
# print(b)
#

# city=['Bangalore','Mumbai','Delhi','Chennai']
# str=""
# for i in city :
#     # str=str+" "+i[::-1]
#     # str1="".join(i[::-1])
#     str=str+" "+"".join(i[::-1])
# print(str)
#
# a={"name":"Prady","age":34,"Company":"DXC","EmployeeType":"Regular"}
#
# # print(a)
#
# print(a['name'])
# print(a.get('name'))
#
# a['name']="Divakar"
#
# a['name']="Prady"
# print(a)
#
# a['car']="Honda Amaze"
#
# print(a)
#
# a['Mileage']=27000
#
# print(a)
#
# a.pop('Mileage')
# print(a)
#
# a['Mileage']=27000
# print(a)
# # b=a.popitem()
# # print(b)
# # b=a.popitem()
# # print(a)

# a={"Employee1":{"name":"Prady","age":34,"Company":"DXC","EmployeeType":"Regular"},
#    "Employee2":{"name":"divakar","age":45,"Company":"DXC","EmployeeType":"Regular"}
# }
#
# a['Employee3']={}
# a['Employee3']['name']="Manoj"
# a['Employee3']['age']=35
# a['Employee3']['Company']="DXC"
# a['Employee3']['EmployeeType']="Regular"

# for key in a.keys():
#     print(key)
#
# for values in a.values():
#     print(values)

# for key,value in a.items():
#     print("Key: ",key)
#     for i_key in value:
#         print(i_key+":",value[i_key])


# Working with Dictionary

# my_dict={"name":"Divakar","age":41,"company":"dxc","sex":"Male"}
# print(my_dict["company"])
# print(my_dict["name"])
#
# print(my_dict.get('name'))
#
# my_dict['location']="Bangalore"
#
# print(my_dict)
#
# my_dict.update({"address":"ranebennur","marital Status":"Married"})
#
# print(my_dict)
#
# b={"vehicle":"jupiter"}
#
# my_dict.update(b)
#
# print(my_dict)
#
# del my_dict['vehicle']
# print(my_dict)

# # # Nested Dictionary
# Emp_Details={"Employee1":{"name":"Prady","age":35,"company":"Dxc"},
#              "Employee2":{"name":"manoj","age":35,"company":"Dxc"},
#
#              }
#
# print(Emp_Details)
#
# Emp_Details["Employee3"]={}
# Emp_Details["Employee3"]["name"]="Divakar"
# Emp_Details["Employee3"]["age"]=41
# Emp_Details["Employee3"]["company"]="Dxc"
#
# Emp_Details["Employee4"]={}
# Emp_Details["Employee4"]["name"]="Karthik"
# Emp_Details["Employee4"]["age"]=36
# Emp_Details["Employee4"]["company"]="Dxc"
#
# print(Emp_Details["Employee2"]["name"])

# Loop over Dictionary
#
# # For iterating keys
# for keys in Emp_Details.keys():
#     print(keys)
#
# # For iterating values
#
# for values in Emp_Details.values():
#     print(values)

# for key,value in Emp_Details.items():
#     print(key+":",value)
#
#
# for dic_key,dic_val in Emp_Details.items():
#     print("Key :",key)
#     for key in dic_val:
#         print(key,dic_val[key])
import datetime
import timeit
#
# lst=[1,2,3,4,5,6,7,8,9]
# start=timeit.default_timer()
# for i in lst:
#     pass
#     # print(i)
# end=timeit.default_timer()
# print("Time",end-start)
#
# tup=(1,2,3,4,5,6,7,8,9)
# start1=timeit.default_timer()
# for i in tup:
#     pass
#     # print(i)
# end1=timeit.default_timer()
# print("Time",end1-start1)


# lst=[1,2,3,4,5,6,7]
# tup=(1,2,3,4,5,6,7)
#
# print(lst.__sizeof__())
# print(tup.__sizeof__())

# Working with functions
#
# def function_name(args):
#    # your code
#    return

# def greet():
#    print("Hi Divakar")
#
#
# greet()
#
# def gree():
#    pass
#
#
# gree()

# def greet(name):
#    print("Hi",name)
#
# greet("Manoj")
# greet("KArthik")

# def get_gender(sex="unknown"):
#    sex=sex.lower()
#    if sex=='m':
#       print("Male")
#    elif sex=='f':
#       print("Female")
#    else:
#       print(sex)
#
# get_gender('m')
# get_gender('f')
# get_gender()
#
# def greet(name,company="DXC"):
#    # print("Hi "+name+",working in "+company)
#       print("Hi {},working in {}".format(name,company))
#
# greet()
#
# Function with *args ( VAriable length arguments )
# def greet(*names):
#    print(type(names))
#    for name in names:
#       print("Hi",name)
#
# # names="Prady","Divakar","Manoj","Karthik"
# greet("Prady","Divakar","Manoj","Karthik")

# def add(*num):
#    sum=0
#    for i in num:
#       sum=sum+i
#    return sum
#
# total=add(1,2,3,4,5,6,7,8,9)
# print(total)

# **kwargs

# def resume(**kwargs):

# def introduction(**data):
#    for key,value in data.items():
#       print(key,value)
#
# introduction(name="Divakar",company="DXC",location="Bangalore",hobbies="PArtying")

# def student(name,age,**marks):
#    print(name)
#    print(age)
#    for subject,mark in marks.items():
#       print(subject+":",mark)
#
# student("PRady",34,english=75,maths=65,social=71,hindi=75)

# d1={1:"name",2:"age"}
# d2={3:"school",4:"result"}
#
# d3={**d1,**d2}
# # d3=d1,d2
# print(d3)

# sentence="look into my eyes and say i love you look into my eyes what are you doing look into my eyes"
# words=sentence.split()
# word_dict={}
# for word in words:
#    word_count=words.count(word)
#    word_dict.update({word:word_count})
#
# print(word_dict)
#
# import time
# import sys
# #
# # str="hello world"
# # for char in str:
# #     print(char,end='')
# #     time.sleep(.25)
#
# str="hello world"
#
# for c in str:
#     sys.stdout.write(c)
#     sys.stdout.flush()
#     time.sleep(.25)

import time
# from Module1.test import add,sub,linear_search


# print(add(5,5))
# print(sub(10,5))
#
#
# lst=[2,3,4,9,1,11,5]
# lst.sort()
# index=linear_search(lst,4)
# if index >0:
#     print("number found in",index)
# else:
#     print("number not found")

# from Module1 import test
#
# print(test.sub(10,5))
# print(test.add(5,5))
# print(test.linear_search([1,2,3,4,5,6],4))

# from Module1.test import *
#
# from Module1.test import add as a
# print(a(5,5))

from Module1.test import add as a
#
# import sys
#
# print(sys.path.append("C:/Users/pr57/Desktop/Selenium_Python"))
#
# print(sys.path)
#
# import Maths as t
# print(t.add(5,5))
#
# # import sys
# # sys.path.append("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files")
# #
# # import PythonModules as PM
# #
# # PM.printforward(10)

## Object Oriented PRogramming

# a=10

# each and every thing is treated as object . 10 is a object with a memory location allocated and x is a reference to it


# name="PRady"
# print(isinstance(name,str))

#
# name="Pradyumna"
# for char in name:
#     print(char,end="")
#     time.sleep(.25)


#
# class Human():
#     pass
# Divakar=Human()
#
#
# #instance Varibales
# Divakar.name="Divakar"
# Divakar.age=39
# Divakar.lastname="CS"
#
# manoj=Human()
# manoj.name="Manoj"
# manoj.age=35
# manoj.lastname="Hande"
#
# print(Divakar.name)
# print(manoj.name)

#
# class Person():
#
#     #Constructor
#     def __init__(self,firstname,lastname,age):
#         self.firstname=firstname
#         self.lastname=lastname
#         self.age=age
#
#     def show_firstname(self):
#         print(self.firstname)
#
#     def show_lastname(self):
#         print(self.lastname)
#
#     def show_age(self):
#         print(self.age)
#
# p1=Person("manoj","hande",35)
# p2=Person("Divakar","CS",39)
#
# p1.show_firstname()
# p1.show_lastname()
# p1.show_age()
#
# p2.show_firstname()
# p2.show_lastname()
# p2.show_age()
#
# class maths():
#
#     # def __init__(self,num1,num2):
#     #     self.num1=num1
#     #     self.num2=num2
#
#     def add(self,num1,num2):
#         print(num1+num2)
#
#     def sub(self,num1,num2):
#         print(num1-num2)
#
#     def mul(self,num1,num2):
#         print(num1*num2)
#
#     def div(self,num1,num2):
#         print(num1%num2)
#
# # m1=maths()
# # m1.add(10,5)
# # m1.sub(44,5)
# # m1.div(25,5)
# # m1.mul(44,55)
#
#
# class Employee():
#
#     #Class Variable
#     salary_raise=1.05
#     company="DXC"
#
#     def __init__(self,name,salary,age):
#         self.name=name
#         self.salary=salary
#         self.age=age
#
#     def Salary_Change(self):
#         print(self.salary*self.salary_raise)
#
#
#
# prady=Employee("Pradyumna",50000,35)
# manoj=Employee("Manoj HAnde",60000,35)
# diva=Employee("Divakar CS",75000,39)
#
# prady.Salary_Change()
# manoj.Salary_Change()
# diva.Salary_Change()
#
#
# Employee.salary_raise=1.07
# # Employee.salary_raise=1.07
#
# prady.Salary_Change()
#
# manoj.Salary_Change()
#
# print(Employee.salary_raise)
# print(Employee.company)

#Generator Function

# def gen_x(str):
#     length=len(str)
#     for i in range(length-1,-1,-1):
#         yield str[i]

# for i in gen_x([1,2,3,4,5,6]):
#     print(i)
# a=gen_x("hello")
# print(next(a))
#
# print(next(a))
#
# def gen_x():
#     n=1
#     yield n
#
#     n=n+1
#     yield n
#
#     n=n+2
#     yield n
#
# a=gen_x()
#
# print(next(a))
# print(next(a))
# print(next(a))

# inheritance

# Inheritance is " is a relationship"
#
# class Person():
#
#     company=None
#
#     def __init__(self,firstname,lastname):
#         self.firstname=firstname
#         self.lastname=lastname
#
#     def showname(self):
#         print('Hi{} {},Please welcome to the class'.format(self.firstname,self.lastname))
#
# class Employee(Person):
#
#     def __init__(self,firstname,lastname,year):
#         Person.__init__(self,firstname,lastname)
#         # super().__init__(firstname,lastname)
#         self.year=year
#
#     def showyear(self):
#         print(self.year)
#
# e1=Employee("Prady","Rama",2007)
# e1.showyear()
# e1.showname()
#
# # e1.company="DXC"
#
# class Polygon():
#
#     height=None
#     width=None
#
#     def setvalues(self,height,width):
#         self.height=height
#         self.width=width
#
# class Triangle(Polygon):
#
#     def area(self):
#         print(self.height*self.width/2)
#
# class Rectangle(Polygon):
#
#     def area(self):
#         print(self.height*self.width)
#
# tri=Triangle()
# tri.setvalues(10,5)
# tri.area()
#
# rect=Rectangle()
# rect.setvalues(15,10)
# rect.area()
#
#
#
# class Polygon():
#
#     def __init__(self,height,width):
#         self.height=height
#         self.width=width
#
# class Triangle(Polygon):
#
#     def area(self):
#         print(self.width*self.height/2)
#
#
# tri=Triangle(5,6)
# tri.area()
#
# class Person():
#
#     def __init__(self,firstname,lastname):
#         self.firstname=firstname
#         self.lastname=lastname
#
#     def showname(self):
#         print('First nameis {},lastname is {}'.format(self.firstname,self.lastname))
#
# class Employee():
#
#     def __init__(self,EID):
#         self.EID=EID
#
#     def showID(self):
#         print("Employee ID is",self.EID)
#
# class Address(Person,Employee):
#
#     def __init__(self,firstname,lastname,EID,location):
#         Person.__init__(self,firstname,lastname)
#         Employee.__init__(self,EID)
#         self.location=location
#
#     def showlocation(self):
#         print("Location is",self.location)
#
# e1=Address("Pradyuma","Ramachandra",82051837,"Bangalore")
# e1.showname()
# e1.showID()
# e1.showlocation()

#Avoiding Conflicts in Multiple inheritance
#
# class Person():
#     def __init__(self):
#         self.firstname="Pradyumna"
#
#     def showname(self):
#         print(self.firstname)
#
# class Employee():
#      def __init__(self):
#          self.firstname="Ramachandra"
#
#      def showname(self):
#          print(self.name)
#
# class Pradyumna(Employee,Person):
#     def __init__(self):
#         super().__init__()
#         # Person.__init__(self)
#         # Employee.__init__(self)
#
#     def showname(self):
#         print(self.firstname)
# #
# p1=Pradyumna()
# # p1.showname()
#
# print(Pradyumna.__mro__)
#
# class Polygon():
#     height=None
#     width=None
#
#     def area(self):
#         print("Polygon Class")
#
#
# class Sides(Polygon):
#
#     def setvalues(self,height,width):
#         self.height=height
#         self.width=width
#
#     def area(self):
#         print("Sides Class")
#         super().area()
#
# class Rectangle(Sides):
#
#     def __init__(self):
#         print("Initializing class Rectangle")
#
#     def area(self):
#         print(self.height*self.width)
#         super().area()
#
#
# rect=Rectangle()
# rect.setvalues(5,8)
# rect.area()

# @classmethod, static method
#
# class Employee():
#
#     raise_amount=1.05
#     bonus=5000
#
#     def __init__(self,firstname,lastname,pay):
#         self.firstname=firstname
#         self.lastname=lastname
#         self.pay=pay
#
#     def showname(self):
#         return f'{self.firstname} {self.lastname}'
#
#     def total_salary(self):
#         print(self.pay*self.raise_amount)
#
#     @classmethod
#     def set_raise_amt(cls,amount):
#         cls.raise_amount=amount
#
#     @classmethod
#     def emp_str(cls,empstr):
#         first,last,pay=empstr.split("-")
#         return cls(first,last,pay)
#
#
#     @staticmethod
#     def is_workday(day):
#         if day.weekday()==5 or day.weekday()==6:
#             return False
#         return True
#
# #
# # e1=Employee("Pradyumna","Ramachandra",75000)
# # # print(e1.showname())
# # e1.total_salary()
# # Employee.set_raise_amt(1.06)
# # e1.total_salary()
# #
# # my_date=datetime.date(2020,5,30)
# # print(my_date)
# # print(Employee.is_workday(my_date))
#
#
# emp_str="Sandhya-Chandrasekhara-75000"
# # first,last,pay=emp_str.split("-")
# e2=Employee.emp_str(emp_str)
# print(e2.showname())

# @property for validating getter and setter
#
# class Person():
#     def __init__(self):
#         self._age=0
#
#     @property
#     def age(self):
#         print("Running getter")
#         return self._age
#
#     @age.setter
#     def age(self,a):
#         if a<18:
#             raise ValueError("Age is less than 18")
#         print("setter method called")
#         self._age=a
#
# mark=Person()
# mark.age=19
# print(mark.age)
#
# class Employee():
#
#     def __init__(self,first,last):
#         self.first=first
#         self.last=last
#         # self.email=first+'.'+last+'@email.com'
#
#     @property
#     def email(self):
#         return self.first+'.'+self.last+'@email.com'
#
#     @property
#     def fullname(self):
#         return self.first+" "+self.last
#
#     @fullname.setter
#     def fullname(self,name):
#         first,last=name.split(" ")
#         self.first=first
#         self.last=last
#
#
#
# e1=Employee("Pradyumna","Ramachandra")
# print(e1.fullname)
# print(e1.email)
# e1.fullname="Sandhya Chandrasekhar"
# print(e1.email)


# Multiple Inheritance is basically where a class will have more than one parent
#
# class Person():
#
#     def __init__(self,firstname,lastname):
#         self.firstname=firstname
#         self.lastname=lastname
#
#     def showname(self):
#         print(self.firstname+" ",self.lastname)
#
# class Employee():
#
#     def __init__(self,employeeid):
#         self.employeeid=employeeid
#
#     def showID(self):
#         print(self.employeeid)
#
#
# class Location(Person,Employee):
#
#     def __init__(self,firstname,lastname,employeeid,location):
#         Person.__init__(self,firstname,lastname)
#         Employee.__init__(self,employeeid)
#         self.location=location
#
#     def showlocation(self):
#         print(self.location)
#
#
# e1=Location("Divakar","CS",2125789,"Bangalore")
# e1.showname()
# e1.showID()
# e1.showlocation()
#
# class Person():
#     def __init__(self):
#         self.name="Divakar"
#
# class Employee():
#     def __init__(self):
#         self.name="Manoj"
#
#
# class C(Person,Employee):
#     def __init__(self):
#         super().__init__()
#
# e1=C()
# print(e1.name)
#
#
# class A:pass
# class B(A):pass
# class C(A):pass
# class D(A):pass
#
# class E(B,C,D):pass
#
# # c3 Alogorithm
#
#       E+A+A+A

#     # E+B+C+D+A+O
#
# print(E.mro())

#
# def outer_func():
#     message="Hello"
#     def inner_func():
#         print(message)
#     inner_func()
#
# outer_func()

# def outer_func(msg):
#     def inner_func():
#         print(msg)
#     return inner_func
#
# a=outer_func("hello")
# a()
#
#
# def outer_func(origfunc):
#     def innerfunc(*args,**kwargs):
#         print("printing {}".format(origfunc.__name__))
#         print(origfunc(*args,**kwargs))
#         print("_____________")
#     return innerfunc
# #
# # add1=outer_func(add)
# # add1(6,5)
#
# @outer_func
# def add(x,y):
#     return x+y
#
# add(6,5)
#
# class outer_func():
#
#     def __init__(self,orignalfunction):
#         self.originalfunction=orignalfunction
#
#     def __call__(self, *args, **kwargs):
#         print("Sleeping for one second")
#         return self.originalfunction(*args, **kwargs)
# #
# # def outer_func(origfunc):
# #     def innerfunc(*args,**kwargs):
# #         print("***********************************")
# #         print("printing {}".format(origfunc.__name__))
# #         return origfunc(*args,**kwargs)
# #
# #     return innerfunc
#
# @outer_func
# def display_info(name,age):
#     print("Display info ran with {} {}".format(name,age))
#
# display_info("Prady",34)
#
# a=[1,2,3,4,5]
# myiter=iter(a)
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
#
# class my_numbers():
#
#     def __iter__(self):
#         self.a=1
#         return self
#
#     def __next__(self):
#         x=self.a
#         if x<=25:
#             self.a+=1
#             return x
#         else:
#             raise StopIteration
# a=my_numbers()
# for i in a :
#     print(i)

#
# string="HEllo"
# def genrator_func(string):
#
#     length=len(string)
#     for i in range(length-1,-1,-1):
#         yield string[i]
#
#
#
# a=genrator_func(string)
# print(next(a))
# print(next(a))
# print(next(a))
# print(next(a))
#
# l=[x*x for x in range(100000000000000000)]
# print(l)

# l=(x*x for x in range(100))
# while True:
#     print(next(l))
#
# a=["Pradyumna Ramachandra","Sandhya Chandrasekhar","Krishnamurthy Ramachandra"]
#
# def sorted_function(lst):
#     mydict={}
#     for i in range(len(lst)):
#         new_list=lst[i].split()
#         mydict[new_list[0]]=new_list[1]
#     return mydict
#
# mydict=sorted_function(a)
# for key,value in sorted(mydict.items()):
#     print(key,value)

#
# words = ['banana', 'pie', 'Washington', 'book']
#
# def sorted_func(word):
#     return word[::-1]
#
# print(sorted(words,key=sorted_func))



# a=["Pradyumna Ramachandra","Sandhya Chandrasekhar","Krishnamurthy Ramachandra"]
#
# def sort_iterable(lst):
#     for i in range(len(lst)):
#         return lst[i].split()[0]
#
# print(sorted(a,key=sort_iterable))
#
# print(sorted(a,key=lambda x:x.split()[0]))
#
# class X:pass
# class Y:pass
# class Z:pass
# class A(X,Y):pass
# class B(Y,Z):pass
# class M(A,B,Z):pass
#
#
# # M+A+X+B+Y+Z+O
# print(M.mro())
#
# string="ABCD1234ABC12"
# mydict={}
# for i in string:
#     char_count=string.count(i)
#     mydict[i]=char_count
#
# print(mydict)

# a=["Pradyumna Ramachandra","Sandhya Chandrasekhar","Krishnamurthy Ramachandra"]

# print(sorted(a,key=lambda x:x.split()[0][::-1]))
#
# def upper_case(func):
#     def inner():
#         x=func()
#         return x.upper()
#     return inner
#
#
# @upper_case
# def display_info():
#     return "good morning"
#
# print(display_info())
#
# def sorted_list(func):
#     def inner(*args):
#         a=func(*args)
#         new_list=sorted(a)
#         return new_list
#     return inner
#
#
# @sorted_list
# def display_lst(lst):
#     return lst
#
# # lst=[8,9,5,3,1,2]
# print(display_lst([8,9,5,3,1,2]))
#
# lst=[8,9,5,[5,3,1,99],3,1,2,[103]]
# flat_list = []
# def max_lst(lst):
#
#     for i in lst:
#         if type(i)==list:
#             max_lst(i)
#         else:
#             flat_list.append(i)
#     return max(flat_list)
#
# print(max_lst(lst))
#
# import datetime
#
# a=datetime.datetime(2020,5,28)
# print(a.weekday()==5)
#
# a=[8,9,5,3,1,2]
# n=len(a)
# for i in range(n):
#     for j in range(n-i-1):
#         if a[j]>a[j+1]:
#             a[j],a[j+1]=a[j+1],a[j]
#
# print(a)
#
#
# phrases = ['when in rome','what goes around comes around','all is fair in love and war']
#
# def custom_sort(a):
#     x=a.split()[2]
#     return x
#
# print(sorted(phrases,key=custom_sort))
#
# n=5
# for i in range(0,n):
#     for j in range(0,i+1):
#         print("*",end=" ")
#     print()
#
# n=5
# k=2*n-2
# for i in range(0,n):
#     for j in range(k):
#         print(end=" ")
#     k=k-1
#     for j in range(0,i+1):
#         print("*",end=" ")
#     print()
#
# n=5
# num=1
# for i in range(0,n):
#     num=1
#     for j in range(0,i+1):
#         print(num,end=" ")
#         num=num+1
#     print()
#
# def upper_case(origfunc):
#     def inner():
#         x=origfunc()
#         x_list=x.split(" ")
#         return x_list
#     return inner
#
#
# @upper_case
# def display():
#     return "Good Morning"
#
# # d=upper_case(display)
# # print(d())
# print(display())

# Multiple Inheritance is basically where a class will have more than one parent
#
# class Person():
#
#     def __init__(self,firstname,lastname):
#         self.firstname=firstname
#         self.lastname=lastname
#
#     def showname(self):
#         print(self.firstname+" ",self.lastname)
#
# class Employee():
#
#     def __init__(self,employeeid):
#         self.employeeid=employeeid
#
#     def showID(self):
#         print(self.employeeid)
#
#
# class Location(Person,Employee):
#
#     def __init__(self,firstname,lastname,employeeid,location):
#         Person.__init__(self,firstname,lastname)
#         Employee.__init__(self,employeeid)
#         self.location=location
#
#     def showlocation(self):
#         print(self.location)
#
#
# e1=Location("Divakar","CS",2125789,"Bangalore")
# e1.showname()
# e1.showID()
# e1.showlocation()

## Multilevel Inheritance
#
# class A:
#     pass
#
# class B(A):
#     pass
#
# class C(B):
#     pass
#
# class Polygon():
#
#     width=None
#     height=None
#
#     def display_info(self):
#         print("Polygon Class")
#
# class set_values(Polygon):
#
#     def set_value(self,width,height):
#         self.width=width
#         self.height=height
#
#     def display_info(self):
#         super().display_info()
#         print("Calling Set values class")
#         print("Calculating Area-------->")
#
# class Rectangle(set_values):
#
#     def area(self):
#         super().display_info()
#         print("Area of Rectangle is ",self.height*self.width)
#
# r=Rectangle()
# r.set_value(6,3)
# r.area()
#
# class Employee():
#
#     #Class Variable
#     raise_amount=1.05
#
#     def __init__(self,firstname,lastname,pay):
#         self.firstname=firstname
#         self.lastname=lastname
#         self.pay=pay
#
#     def display_name(self):
#         print(self.firstname+" ",self.lastname)
#
#     @classmethod
#     def set_raise_amount(cls,amount):
#         cls.raise_amount=amount
#
#     @staticmethod
#     def total_Salary(x,y):
#         print(x+y)
#
#     @staticmethod
#     def is_workday(emp_date):
#         if emp_date.weekday()==5 or emp_date.weekday()==6:
#             return False
#         return True
#
#
#
# e1=Employee("Pradyumna","Ramachandra",50000)
# # e1.raise_amount=2.05
# print(e1.raise_amount)
#
# # print(e1.raise_amount)
# #
# Employee.set_raise_amount(3.05)
# # print(Employee.raise_amount)
# #
# # print(e1.raise_amount)
#
# e2=Employee("xxxxx","yyyyy",50000)
# print(e2.raise_amount)
#
# Employee.total_Salary(10,15)
#
# import datetime
#
# my_date=datetime.date(2020,6,6)
# print(Employee.is_workday(my_date))
#
#
#
#
#
#
# # MEthod Overloading  same class not possible, it will overide
# class Maths():
#
#     def add(self,x,y):
#         return x+y
#
#     def add(self,x,y,z):
#         return x+y+z
#
# m=Maths()
# print(m.add(5,6,7))
#
# #Method Overloading
# class display_info():
#
#     def show_name(self,name=None):
#         if name is not None:
#             print(name)
#         else:
#             print("Name is none")
#
# #Decorators Class
# class upper_case():
#
#     def __init__(self,orignalfunc):
#         self.originalfunc=orignalfunc
#
#     def __call__(self, *args, **kwargs):
#         print("Printing {}".format(self.originalfunc.__name__))
#         x=self.originalfunc()
#         return x.upper()
#
# @upper_case
# def display_info():
#     return "Good Morning"
#
#
# print(display_info())


# x=["Pradyumna Ramchandra","Sandhya Chandra","Samrudh PRadyumna"]
#
# def custom_sort(a):
#     return a.split()[1]
#
# print(sorted(x,key=custom_sort))
#
#
# x=["i am god","god is great","war and game"]
#
# def custom_sort(a):
#     return a.split()[2]
#
# print(sorted(x,key=custom_sort))


# @property for validating getter and setter

# class Person():
#     def __init__(self):
#         self._age=0
#
#     @property
#     def age(self):
#         print("Running getter")
#         return self._age
#
#     @age.setter
#     def age(self,a):
#         if a<18:
#             raise ValueError("Age is less than 18")
#         print("setter method called")
#         self._age=a
#
# mark=Person()
# mark.age=19
# print(mark.age)
#
#
# # @property getter and setter
#
# class Employee():
#     def __init__(self,first,last,pay):
#         self.first=first
#         self.last=last
#         self.pay=pay
#
#     @property
#     def email(self):
#         return '{}.{}@email.com'.format(self.first,self.last)
#
#     @property
#     def fullname(self):
#         return '{} {}'.format(self.first,self.last)
#
#     @fullname.setter
#     def fullname(self,name):
#         first, last=name.split(' ')
#         self.first=first
#         self.last=last
#
#     @fullname.deleter
#     def fullname(self):
#         print('Deleting Name !!!')
#         self.first = None
#         self.last = None
#
#
# e1=Employee("Pradyumna","Ramachandra",5000)
# print(e1.fullname)
# e1.fullname="Sandhya Chandra"
# print(e1.fullname)
# del e1.fullname

#
# class Car():
#
#     __ownername="Pradyumna"
#
#     @property
#     def get_owner_name(self):
#         print("Getter method called")
#         return self.__ownername
#
#     @get_owner_name.setter
#     def get_owner_name(self,name):
#         print("Setter method called")
#         self.__ownername=name
#
# c=Car()
# c.get_owner_name="Sandhya"
# print(c.get_owner_name)

# Encapsulation

# Dunders   __variablename
#
# from abc import ABC,abstractmethod
#
#
# class A(ABC):
# __methodname

#
# class Student():
#
#     __marks=500
#
#     def __init__(self,name):
#         self.name=name
#         # self.__marks=500
#
#     def set_marks(self,value):
#         self.__marks=value
#
#     def get_marks(self):
#         # return self.__marks
#         self.__showmarks()
#
#     # PRivate method accessing the private variable
#     def __showmarks(self):
#         print(self.__marks)
#
# s1=Student("MAnoj")
# s1.get_marks()
# #
# # print(s1.name)
# # # print(s1.__marks)
# #
# # s1.set_marks(550)
# # print(s1.get_marks())
# #
# # s1.__marks=600
# #
# # print(s1.get_marks())
# #
# # ##
# # print(Student._Student__marks)
#
# class Student():
#
#     def __init__(self,name,grade):
#         self.name=name
#         self.grade=grade
#         # self.msg=self.name+" has got "+self.grade
#
#     @property
#     def msg(self):
#         return self.name+" has got Grade"+self.grade
#
#     @msg.setter
#     def msg(self,msg_name):
#         a,*b,c=msg_name.split(" ")
#         self.name=a
#         self.grade=c
#
#
#
#
# s1=Student("Prady","A")
# print(s1.msg)
# s1.grade="B"
# print(s1.msg)
# s1.msg="Amulya has got Grade A+"
# print(s1.msg)
#
#
# d="Amulya has got Grade A+"
# a,*b,c=d.split(" ")
#
# print(a)
# print(b)
# print(c)

## File Handling

#
# fh=open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\log3.txt","r")
#
# for i in fh:
#     print(i)
# # # print(type(fh.read()))
# # content=fh.read()
# # for char in content:
# #     print(char)
#
#
# fh.close()


# with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\log3.txt","r") as fh:
    # content=fh.readline()
    # print(fh.readline(5))
    # print(fh.read(10))
    # print(fh.readline(10))
    # print(fh.readline(5))
    # fh.seek(22)
    # print(fh.readline())
    #  print(type(fh.readline()))

# with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\log3.txt", "r") as fh:
#     content=fh.readlines()
#     for line in content:
#         print(line,end="")

# with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\log3.txt", "r") as fh:
#     content=fh.readlines()
#     with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\log4.txt", "a") as fw:
#         fw.writelines(content)
#         # for line in fh:
#         #     fw.write(line)
#
# with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\log4.txt", "w") as fw:
#     for i in range(26):
#         fw.write("This is line number %d\n"%(i+1))

#
#
# x=(i for i in range(100))
# while True:
#     print(next(x))
#
# string="Hello"
#
# def rev_string(a):
#     for char in range(len(a)-1,-1,-1):
#         yield a[char]
#
# b=rev_string(string)
#
# print(next(b))
# print(next(b))
# print(next(b))
# print(next(b))
# print(next(b))
# print(next(b))

# Exception Handling
#
# try:
#     # exception code
# except:
#     to print the exception

# num1=int(input("Enter the number 1 :"))
# num2=int(input("Enter the number 2 :"))
#
# try:
#     result=num1/num2
#     print(result)
# except Exception as e:
#     print("The Exception is ",e)

#
# string="Pradyumna"
# try:
#     # print(5/0)
#     print(string[9])
# except Exception as e:
#     print("The Exception",e)
# except ArithmeticError as e:
#     print(e)
# except IndexError as e:
#     print(e)
#
# string="Pradyumna"
# # print(string[9])
# try:
#     print(string[9])
# except Exception as e:
#     print(e)
# # else:
#     print("Else Block is being executed")
# finally:
#     print("Finally Block being executed")



# print("Divakar")

# Raising Exceptions
#
# string="Pradyumna"
# index=5
# if index >8:
#     raise Exception("The len of the sting is 8 ")
# print(string[index])

#
# class CoffeeCup():
#     def __init__(self,temp):
#         self.temp=temp
#
#     def drink_coffee(self):
#         if self.temp>85:
#             raise Exception("The coffee is too hot")
#         elif self.temp<65:
#             raise Exception("The coffee is too cold")
#         else:
#             print("Enjoy your coffee")
#
# c1=CoffeeCup(64)
# c1.drink_coffee()
#
#
#
#
#
# import json

# person={"Firstname":"Divakar",
#         "Lastname":"CS",
#         "marks":[65,76,66,88,99],
#         "result":"PASS"
# }
#
# # a=json.dumps(person)
#
# with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\sample.json","w") as fh:
#     fh.write(json.dumps(person,indent=2))
#
# with open("C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\example_2.json","r") as fh:
#     json_value=json.load(fh)
#     d=json_value['quiz']['maths']
#     for key,value in d.items():
#         for d_key in value:
#             print(d_key,value[d_key])

#
#


# n=5
# for i in range(n):
#     for j in range(i+1):
#         print("*",end="")
#     print()
#
# n=7
# k=2*n-2
# for i in range(n):
#     for j in range(k):
#         print(end=" ")
#     k=k-1
#     for j in range(i+1):
#         print("*",end=" ")
#     print()
#
# n=5
# num=1
# for i in range(n):
#     num=1
#     for j in range(i+1):
#         print(num,end=" ")
#         num+=1
#     print()

#
# def upper_case(func):
#     def inner():
#         x=func()
#         a=x.upper()
#         print(type(a))
#         return a
#     return inner
#
# def str_split(func):
#     def inner():
#         a=func()
#         print(type(a))
#         return a.split()
#     return inner
#
# @str_split
# @upper_case
# def msg():
#     return "good morning"
#
# def first_dec(func):
#     def inner():
#         return "first "+func()+" first"
#     return inner
#
# def second_dec(func):
#     def wrapper():
#         return "second "+func()+" second"
#     return wrapper
#
# @second_dec
# @first_dec
# def msg():
#     return "good morning"
#
# print(msg())
#
# a=1
# b=2
# n=10
# while n-2:
#     c=a+b
#     a=b
#     b=c
#     print(c,end=" ")
#     n=n-1

#
# num=11
# if num>1:
#     for i in range(2,num):
#         if num%i==0:
#             print("Not a Prime Number")
#             break
#     else:
#         print("Prime Number")
# else:
#     print("Not a Prime Number")
#
# a=[7,5,3]
# n=len(a)
# for i in range(n):
#     for j in range(n-i-1):
#         if a[j]>a[j+1]:
#             a[j],a[j+1]=a[j+1],a[j]
#
# print(a)

#
# a=[1,2,3,4,5,6,7,8,9]
# key=7
# start=0
# end=len(a)
# while start<end:
#     mid=(start+end)//2
#     if a[mid]>key:
#         end=mid
#     elif a[mid]<key:
#         start=mid+1
#     else:
#         print(mid)
#         break
#
# str1="hello"
# def gen_func(str1):
#     for char in range(len(str1)-1,-1,-1):
#         yield str1[char]
#
# a=gen_func(str1)
# print(next(a))
# print(next(a))
# print(next(a))
# print(next(a))
# print(next(a))
# print(next(a))
#
# import timeit
# start=timeit.default_timer()
# time.sleep(2)
# end=timeit.default_timer()
# print(end-start)

# a=(i*2 for i in range(10000000000))
# while True:
#     print(next(a))
#
#
# # l=[1,2,3,4,5,6,7,8,9,10,11]
# t=(1,2,3,4,5,6,7,8,9,10,11)
# #
# # print(l.__sizeof__())
# # print(t.__sizeof__())
#
# start=timeit.default_timer()
# for i in t:
#     pass
# end=timeit.default_timer()
# print(end-start)
#
# str1=["Pradyumna Ramachandra","Manoja Hande","Divakara Shimog"]
#
# str2=sorted(str1,key=lambda x:x.split()[1][::-1])
# print(str2)

#
# a = {'a': 4, 'b': 3, 'c': 2, 'd': 1}
# print(sorted(a.items(),key=lambda x:x[1]))

import openpyxl
#
# path="C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\TestData.xlsx"
#
# workbook=openpyxl.load_workbook(path)
# worksheet=workbook['Sheet1']
# rows=worksheet.max_row
# cols=worksheet.max_column
# print(rows)
# print(cols)
# #
# # for row in range(1,rows):
# #     for col in range(1,cols):
# #         print(worksheet.cell(row=row,column=col).value,end=" ")
# #     print()
# for row in range(2,rows+1):
#     for col in range(1,cols+1):
#         worksheet.cell(row=row,column=col).value=col
#
#
# workbook.save(path)
#
# import openpyxl
#
# path="C:\\Users\\pr57\\Desktop\\SSM\\Python\\PythonCourse\\Files\\TestData.xlsx"
# workbook=openpyxl.load_workbook(path)
# worksheet=workbook['Sheet1']
#
# # print(worksheet.cell(row=3,column=3).value)
# # print(worksheet['A1'].value)
#
# rows=worksheet.max_row
# print(rows)
# cols=worksheet.max_column
# #
# # dict2=[]
# # for row in range(1,rows+1):
# #     my_dict = {}
# #     for col in range(1,cols+1):
# #         # my_dict.update({})
# #         my_dict[worksheet.cell(row=1,column=col).value]=worksheet.cell(row=row+1,column=col).value
# #         # cellvalue=worksheet.cell(row=row,column=col).value
# #         # print(cellvalue,end=" ")
# #     dict2.append(my_dict)
# #
# #
# # print(dict2)
# for row in range(2,rows+1):
#     for col in range(1,cols+1):
#         worksheet.cell(row=row,column=col).value="TC0"+str(row-1)
#
# workbook.save(path)
